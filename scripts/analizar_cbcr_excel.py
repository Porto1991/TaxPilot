#!/usr/bin/env python3
"""Chequeo estructural de un Excel con datos public CbCR.

Contrasta el contenido contra los elementos exigidos por el art. 48 quater,
ap. 2, de la Directiva 2013/34/UE (introducido por la Directiva (UE) 2021/2101).

Uso:
    python3 scripts/analizar_cbcr_excel.py clientes/fichero.xlsx
"""
import sys
import unicodedata

from openpyxl import load_workbook

# Elementos del art. 48 quater, ap. 2 que deben aparecer desglosados por
# territorio fiscal, con sinónimos habituales (ES/EN) para localizar columnas.
CAMPOS_REQUERIDOS = {
    "territorio fiscal": ["territorio", "jurisdiccion", "jurisdiction", "pais", "country", "tax jurisdiction"],
    "empleados (FTE)": ["empleados", "employees", "fte", "plantilla", "headcount"],
    "ingresos [art. 48 quater 2.d]": ["ingresos", "revenues", "revenue", "turnover", "volumen de negocios", "cifra de negocios"],
    "beneficio/perdida antes de IS [2.e]": ["beneficio", "perdida", "profit", "loss", "pbt", "resultado antes de impuestos", "profit before tax"],
    "IS devengado [2.f]": ["devengado", "accrued", "tax accrued", "impuesto devengado", "current tax"],
    "IS pagado [2.g]": ["pagado", "paid", "tax paid", "impuesto pagado", "abonado", "cash tax"],
    "reservas [2.h]": ["reservas", "accumulated earnings", "retained earnings", "beneficios no distribuidos"],
}

# Estados miembros UE-27: el art. 48 quater, ap. 5 exige desglose individual
# por cada uno de ellos (nunca agregados).
ESTADOS_MIEMBROS = [
    "alemania", "austria", "belgica", "bulgaria", "chipre", "croacia",
    "chequia", "republica checa", "dinamarca", "eslovaquia", "eslovenia",
    "espana", "estonia", "finlandia", "francia", "grecia", "hungria",
    "irlanda", "italia", "letonia", "lituania", "luxemburgo", "malta",
    "paises bajos", "polonia", "portugal", "rumania", "suecia",
]


def normaliza(texto):
    if texto is None:
        return ""
    s = str(texto).strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def busca_cabeceras(hoja, max_filas=15):
    """Localiza la fila de cabeceras: la que más campos requeridos casa."""
    mejor = (0, None, {})
    for fila in hoja.iter_rows(min_row=1, max_row=max_filas):
        encontrados = {}
        for celda in fila:
            texto = normaliza(celda.value)
            if not texto:
                continue
            for campo, sinonimos in CAMPOS_REQUERIDOS.items():
                if campo not in encontrados and any(s in texto for s in sinonimos):
                    encontrados[campo] = (celda.row, celda.column, str(celda.value).strip())
        if len(encontrados) > mejor[0]:
            mejor = (len(encontrados), fila[0].row, encontrados)
    return mejor[1], mejor[2]


def analiza_hoja(hoja):
    fila_cab, encontrados = busca_cabeceras(hoja)
    if not encontrados:
        return None
    faltantes = [c for c in CAMPOS_REQUERIDOS if c not in encontrados]

    jurisdicciones = []
    col_territorio = encontrados.get("territorio fiscal", (None, None, None))[1]
    if col_territorio:
        for fila in hoja.iter_rows(min_row=fila_cab + 1, min_col=col_territorio, max_col=col_territorio):
            v = normaliza(fila[0].value)
            if v and v not in ("total", "totales"):
                jurisdicciones.append(v)

    return {
        "hoja": hoja.title,
        "fila_cabeceras": fila_cab,
        "encontrados": encontrados,
        "faltantes": faltantes,
        "jurisdicciones": jurisdicciones,
    }


def main(ruta):
    wb = load_workbook(ruta, data_only=True)
    print(f"Fichero: {ruta}")
    print(f"Hojas: {', '.join(wb.sheetnames)}\n")

    hallazgos = [r for hoja in wb.worksheets if (r := analiza_hoja(hoja))]
    if not hallazgos:
        print("✗ No se ha localizado ninguna hoja con estructura de informe CbCR.")
        print("  Revisar manualmente: puede usar una disposición no tabular.")
        return 1

    for r in hallazgos:
        print(f"── Hoja «{r['hoja']}» (cabeceras en fila {r['fila_cabeceras']}) ──")
        for campo, (f, c, texto) in r["encontrados"].items():
            print(f"  ✓ {campo}: columna {c} («{texto}»)")
        for campo in r["faltantes"]:
            print(f"  ✗ FALTA: {campo} — exigido por el art. 48 quater, ap. 2")

        juris = r["jurisdicciones"]
        if juris:
            ue_presentes = {j for j in juris if j in ESTADOS_MIEMBROS}
            print(f"  · {len(juris)} filas de territorio fiscal; {len(ue_presentes)} Estados miembros UE identificados")
            print("  · Recordatorio [art. 48 quater, ap. 5]: desglose individual obligatorio por cada")
            print("    Estado miembro y por cada territorio de los Anexos I/II de la lista UE de no")
            print("    cooperadores; el resto puede agregarse.")
        print()

    print("Nota: chequeo estructural automático. El análisis de fondo (atribución por")
    print("territorio, coherencia con EEFF consolidados, moneda, plazos) requiere revisión experta.")
    return 0


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(__doc__)
        sys.exit(2)
    sys.exit(main(sys.argv[1]))

#!/usr/bin/env python3
"""Surgical lookup into the public CbCR jurisdiction matrix.

Returns the exact cell(s) for a jurisdiction (and optionally one field),
always together with the cell reference, the source and the verification
status — so every answer is traceable.

Usage:
    python3 scripts/query_matrix.py <jurisdiction> [field words...]

Examples:
    python3 scripts/query_matrix.py Romania
    python3 scripts/query_matrix.py Ireland penalties
    python3 scripts/query_matrix.py Australia threshold
"""
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

MATRIX = "knowledge/pcbcr_jurisdiction_matrix.xlsx"


def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 2
    country = argv[1].lower()
    field_q = " ".join(argv[2:]).lower()

    ws = load_workbook(MATRIX, data_only=True)["Jurisdiction Matrix"]
    headers = {c: ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)}

    row = next((r for r in range(2, ws.max_row + 1)
                if country in str(ws.cell(row=r, column=1).value).lower()), None)
    if row is None:
        known = [str(ws.cell(row=r, column=1).value) for r in range(2, ws.max_row + 1)]
        print(f"Jurisdiction '{argv[1]}' not in the matrix. Available: {', '.join(known)}")
        return 1

    name = ws.cell(row=row, column=1).value
    src_col = next(c for c, h in headers.items() if h == "Source")
    ver_col = next(c for c, h in headers.items() if h == "Verification status")

    print(f"Jurisdiction: {name}  (row {row})")
    print(f"Source:       {ws.cell(row=row, column=src_col).value}")
    print(f"Verification: {ws.cell(row=row, column=ver_col).value}\n")

    cols = [c for c, h in headers.items()
            if c not in (1, src_col, ver_col) and (not field_q or field_q in str(h).lower())]
    if not cols:
        print(f"No field matches '{field_q}'. Fields: " +
              "; ".join(str(h) for h in headers.values()))
        return 1
    for c in cols:
        val = ws.cell(row=row, column=c).value
        print(f"[{get_column_letter(c)}{row}] {headers[c]}:")
        print(f"    {val}\n")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))

#!/usr/bin/env python3
"""Structural check of an Excel file containing public CbCR data.

Checks the content against Art. 48c of Directive 2013/34/EU (as inserted by
Directive (EU) 2021/2101):
  - required fields of Art. 48c(2), per tax jurisdiction;
  - jurisdiction classification under Art. 48c(5): EU Member State / Annex I
    (1 March of the reporting FY) / Annex II (two-year rule) / other;
  - totals consistency (Total row vs sum of jurisdiction rows).

Usage:
    python3 scripts/analyze_cbcr_excel.py clients/file.xlsx [--fy 2025]
"""
import re
import sys
import unicodedata
from datetime import date
from pathlib import Path

import yaml
from openpyxl import load_workbook

REPO = Path(__file__).resolve().parent.parent
SNAPSHOTS = REPO / "corpus" / "eu" / "eu_list_snapshots.yaml"

REQUIRED_FIELDS = {
    "tax jurisdiction": ["jurisdiction", "jurisdiccion", "country", "pais", "territorio", "tax jurisdiction"],
    "employees (FTE)": ["employees", "empleados", "fte", "headcount", "plantilla"],
    "revenues [Art. 48c(2)(d)]": ["revenues", "revenue", "turnover", "ingresos", "volumen de negocios", "cifra de negocios"],
    "profit/loss before income tax [(2)(e)]": ["profit", "loss", "pbt", "beneficio", "perdida", "resultado antes de impuestos", "profit before tax"],
    "income tax accrued [(2)(f)]": ["accrued", "tax accrued", "current tax", "devengado", "impuesto devengado"],
    "income tax paid [(2)(g)]": ["paid", "tax paid", "cash tax", "pagado", "impuesto pagado", "abonado"],
    "accumulated earnings [(2)(h)]": ["accumulated earnings", "retained earnings", "reservas", "beneficios no distribuidos"],
}

MEMBER_STATES = {
    "austria": "Austria", "belgium": "Belgium", "belgica": "Belgium", "bulgaria": "Bulgaria",
    "croatia": "Croatia", "croacia": "Croatia", "cyprus": "Cyprus", "chipre": "Cyprus",
    "czechia": "Czechia", "czech republic": "Czechia", "chequia": "Czechia", "republica checa": "Czechia",
    "denmark": "Denmark", "dinamarca": "Denmark", "estonia": "Estonia", "finland": "Finland",
    "finlandia": "Finland", "france": "France", "francia": "France", "germany": "Germany",
    "alemania": "Germany", "greece": "Greece", "grecia": "Greece", "hungary": "Hungary",
    "hungria": "Hungary", "ireland": "Ireland", "irlanda": "Ireland", "italy": "Italy",
    "italia": "Italy", "latvia": "Latvia", "letonia": "Latvia", "lithuania": "Lithuania",
    "lituania": "Lithuania", "luxembourg": "Luxembourg", "luxemburgo": "Luxembourg",
    "malta": "Malta", "netherlands": "Netherlands", "paises bajos": "Netherlands",
    "poland": "Poland", "polonia": "Poland", "portugal": "Portugal", "romania": "Romania",
    "rumania": "Romania", "slovakia": "Slovakia", "eslovaquia": "Slovakia",
    "slovenia": "Slovenia", "eslovenia": "Slovenia", "spain": "Spain", "espana": "Spain",
    "sweden": "Sweden", "suecia": "Sweden",
}

# Spanish/English aliases for jurisdictions relevant to the EU annexes.
ALIASES = {
    "islas caiman": "Cayman Islands", "cayman islands": "Cayman Islands",
    "islas virgenes britanicas": "British Virgin Islands", "rusia": "Russian Federation",
    "russia": "Russian Federation", "panama": "Panama", "vanuatu": "Vanuatu",
    "guam": "Guam", "samoa americana": "American Samoa", "anguila": "Anguilla",
    "islas turcas y caicos": "Turks and Caicos Islands", "vietnam": "Viet Nam",
    "islas virgenes de los estados unidos": "US Virgin Islands", "belice": "Belize",
    "turquia": "Türkiye", "turkiye": "Türkiye", "marruecos": "Morocco",
    "groenlandia": "Greenland", "jordania": "Jordan", "montenegro": "Montenegro",
    "suiza": "Switzerland", "switzerland": "Switzerland",
}


def normalize(text):
    if text is None:
        return ""
    s = str(text).strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def canonical(name):
    n = normalize(name)
    if n in MEMBER_STATES:
        return MEMBER_STATES[n], "EU_MS"
    if n in ALIASES:
        return ALIASES[n], None
    return str(name).strip(), None


def load_annexes(fy):
    """Annex I at 1 March of FY; Annex II at 1 March of FY and of FY-1 (two-year rule)."""
    data = yaml.safe_load(SNAPSHOTS.read_text())
    updates = sorted(data["updates"], key=lambda u: str(u["adopted"]))

    def in_force(on):
        applicable = [u for u in updates if str(u["adopted"]) <= on.isoformat()]
        return applicable[-1] if applicable else None

    s_now, s_prev = in_force(date(fy, 3, 1)), in_force(date(fy - 1, 3, 1))
    if not s_now or not s_prev:
        return None
    annex1 = set(s_now["annex_i"])
    annex2 = set(s_now["annex_ii"]) & set(s_prev["annex_ii"])
    return {"annex1": annex1, "annex2_two_year": annex2,
            "snapshot_now": str(s_now["adopted"]), "snapshot_prev": str(s_prev["adopted"])}


def find_headers(sheet, max_rows=15):
    best = (0, None, {})
    for row in sheet.iter_rows(min_row=1, max_row=max_rows):
        found = {}
        for cell in row:
            text = normalize(cell.value)
            if not text:
                continue
            for field, synonyms in REQUIRED_FIELDS.items():
                if field not in found and any(s in text for s in synonyms):
                    found[field] = (cell.row, cell.column, str(cell.value).strip())
        if len(found) > best[0]:
            best = (len(found), row[0].row, found)
    return best[1], best[2]


def analyze_sheet(sheet, annexes):
    header_row, found = find_headers(sheet)
    if not found:
        return False
    missing = [f for f in REQUIRED_FIELDS if f not in found]

    print(f"── Sheet '{sheet.title}' (headers in row {header_row}) ──")
    for field, (_, col, text) in found.items():
        print(f"  ✓ {field}: column {col} ('{text}')")
    for field in missing:
        print(f"  ✗ MISSING: {field} — required by Art. 48c(2)")

    juris_col = found.get("tax jurisdiction", (None, None, None))[1]
    numeric_cols = [c for f, (_, c, _) in found.items() if f != "tax jurisdiction"]
    if not juris_col:
        print()
        return True

    rows, total_row = [], None
    for row in sheet.iter_rows(min_row=header_row + 1, min_col=1, max_col=sheet.max_column):
        raw = row[juris_col - 1].value
        n = normalize(raw)
        if not n:
            continue
        if n in ("total", "totals", "totales"):
            total_row = row
        else:
            rows.append(row)

    # Classification under Art. 48c(5)
    print(f"\n  Jurisdiction classification [Art. 48c(5)] — Annex I as of 1 Mar (snapshot "
          f"{annexes['snapshot_now']}), Annex II two-year rule (also {annexes['snapshot_prev']}):")
    for row in rows:
        name, kind = canonical(row[juris_col - 1].value)
        if kind == "EU_MS":
            tag = "EU Member State — separate disclosure mandatory"
        elif name in annexes["annex1"]:
            tag = "⚠ ANNEX I — separate disclosure mandatory, safeguard-clause omission PROHIBITED [Art. 48c(6)]"
        elif name in annexes["annex2_two_year"]:
            tag = "⚠ ANNEX II (two-year rule) — separate disclosure mandatory, omission prohibited"
        else:
            tag = "Other jurisdiction — may be aggregated"
        print(f"    · {name}: {tag}")

    # Totals consistency
    if total_row is not None:
        print("\n  Totals consistency:")
        for c in numeric_cols:
            vals = [r[c - 1].value for r in rows if isinstance(r[c - 1].value, (int, float))]
            declared = total_row[c - 1].value
            if isinstance(declared, (int, float)):
                s = sum(vals)
                ok = abs(s - declared) < 0.5
                mark = "✓" if ok else "✗"
                head = sheet.cell(row=header_row, column=c).value
                msg = f"    {mark} {head}: rows sum to {s:,.0f}, Total row says {declared:,.0f}"
                if not ok:
                    msg += f"  → DISCREPANCY of {declared - s:,.0f}"
                print(msg)
    else:
        print("\n  · No Total row found — totals consistency not checked.")
    print()
    return True


def main(argv):
    if len(argv) < 2:
        print(__doc__)
        return 2
    path = argv[1]
    fy = int(argv[argv.index("--fy") + 1]) if "--fy" in argv else 2025

    annexes = load_annexes(fy)
    if annexes is None:
        print(f"No EU-list snapshot available for FY{fy} in {SNAPSHOTS} — classification skipped.")
        return 1

    wb = load_workbook(path, data_only=True)
    print(f"File: {path}\nReporting FY: {fy}\nSheets: {', '.join(wb.sheetnames)}\n")
    any_found = any(analyze_sheet(ws, annexes) for ws in wb.worksheets)
    if not any_found:
        print("✗ No sheet with a CbCR report structure was found. Review manually.")
        return 1

    print("Note: automated structural check only. Substantive analysis (attribution per")
    print("jurisdiction, consistency with consolidated FS, currency, deadlines) requires expert review.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))

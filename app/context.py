"""Builds the grounding context for TaxPilot's chat from the corpus.

The corpus is small enough to give the model the full authoritative texts
(with prompt caching this is cheap after the first request), which maximises
accuracy versus a retrieval step that could miss the relevant passage.
"""
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

REPO = Path(__file__).resolve().parent.parent


def _pdf_text(path, first=None, last=None):
    reader = PdfReader(path)
    pages = reader.pages[slice(first, last)]
    return "\n".join((p.extract_text() or "") for p in pages)


def _matrix_tsv():
    ws = load_workbook(REPO / "knowledge" / "pcbcr_jurisdiction_matrix.xlsx",
                       data_only=True)["Jurisdiction Matrix"]
    lines = []
    for row in ws.iter_rows(values_only=True):
        lines.append("\t".join("" if v is None else str(v).replace("\t", " ") for v in row))
    return "\n".join(lines)


@lru_cache(maxsize=1)
def grounding_blocks():
    """Returns the corpus as labelled text blocks for the system prompt."""
    return [
        ("CORPUS INDEX (corpus/index.yaml)",
         (REPO / "corpus" / "index.yaml").read_text()),
        ("DIRECTIVE (EU) 2021/2101 — OFFICIAL ENGLISH TEXT (EUR-Lex)",
         _pdf_text(REPO / "corpus" / "eu" / "2021-2101_directive_public_cbcr_EN.pdf")),
        ("SPAIN — LEY 22/2015, DA 11ª AND ART. 3 DEFINITIONS (BOE consolidated text, "
         "official only in Spanish; pp. 21-23 and 80-85 of the PDF)",
         _pdf_text(REPO / "corpus" / "es" / "ley-22-2015_auditoria_consolidada_ES.pdf", 20, 23)
         + "\n[...]\n"
         + _pdf_text(REPO / "corpus" / "es" / "ley-22-2015_auditoria_consolidada_ES.pdf", 79, 85)),
        ("EU LIST OF NON-COOPERATIVE JURISDICTIONS — VERIFIED SNAPSHOTS "
         "(corpus/eu/eu_list_snapshots.yaml)",
         (REPO / "corpus" / "eu" / "eu_list_snapshots.yaml").read_text()),
        ("ICAC CRITERION — BOICAC 144/2025 QUERY 5 (official, Spanish): ultimate parent "
         "resident in another Member State → that State's rules and deadlines govern",
         _pdf_text(REPO / "corpus" / "es" / "boicac144_consulta5_icac_ES.pdf")),
        ("JURISDICTION MATRIX (knowledge/pcbcr_jurisdiction_matrix.xlsx, TSV export; "
         "SECONDARY SOURCE — check the Verification status column)",
         _matrix_tsv()),
    ]


SYSTEM_RULES = """\
You are TaxPilot, an expert international tax assistant specialising in Public \
Country-by-Country Reporting (Directive (EU) 2021/2101 and national transpositions), \
supporting expert consultants at an international tax advisory firm.

Golden rules:
1. Every normative statement must carry a citation to a specific provision, e.g. \
[Directive (EU) 2021/2101, Art. 48b(1)] or [Ley 22/2015, DA 11ª, ap. tercero.1].
2. The corpus blocks below are your ONLY source of normative truth. If a question is \
not covered by them, say so explicitly; you may add general knowledge but flag it: \
"⚠️ Not verified against the corpus — confirm before sending to a client."
3. Data from the JURISDICTION MATRIX is a secondary source: always report its \
verification status alongside the answer. On any conflict the corpus overrides it.
4. Never invent articles, deadlines, thresholds or references. Flag ambiguities.
5. Distinguish the Directive (EU baseline) from national transpositions. If the user \
does not state a jurisdiction, ask or answer under the Directive and say so.
6. Spanish Directive articles 48 bis/ter/quater/quinquies map to 48a/b/c/d in English; \
cite consistently with the language you write in.
7. Reply in the user's language. Structure: short conclusion → reasoned analysis with \
citations → open points/risks. End substantive answers with: "Draft prepared with AI \
assistance. Subject to review by a qualified professional."
"""


@lru_cache(maxsize=1)
def system_prompt():
    parts = [SYSTEM_RULES, "\n\n===== CORPUS =====\n"]
    for title, text in grounding_blocks():
        parts.append(f"\n----- {title} -----\n{text}\n")
    return "".join(parts)
